"""
Excel-Exportfunktionen für das VALEO-NeuroERP-System.

Dieses Modul enthält Celery-Tasks für die asynchrone Generierung von Excel-Exporten.
"""

import os
import logging
import json
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
from celery import shared_task

# Excel-Generierung
# OpenPyXL wird verwendet, da es die flexibelste Python-Bibliothek für die Arbeit mit Excel-Dateien ist.
# Sie unterstützt sowohl das Lesen als auch das Schreiben von Excel-Dateien und bietet umfangreiche
# Formatierungsoptionen. Pandas wird für die Datenmanipulation und einfache Excel-Exporte verwendet,
# während OpenPyXL für komplexere Formatierungen und Diagramme eingesetzt wird.
try:
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("Pandas, NumPy oder OpenPyXL nicht installiert. Excel-Export ist eingeschränkt.")

# Lokale Imports
from backend.services.task_queue import update_task_progress
from .report_utils import format_report_data, validate_report_parameters

logger = logging.getLogger(__name__)

@shared_task(bind=True)
def generate_excel_export(self, export_data: Dict[str, Any], 
                        output_file: Optional[str] = None) -> Dict[str, Any]:
    """
    Generiert einen Excel-Export basierend auf den übergebenen Daten.
    
    Args:
        export_data: Dictionary mit den Daten für den Export
        output_file: Pfad zur Ausgabedatei (optional)
    
    Returns:
        Dict mit Informationen zum generierten Export
    """
    logger.info("Starte Excel-Export")
    
    try:
        if not EXCEL_AVAILABLE:
            raise ImportError("Pandas, NumPy oder OpenPyXL ist nicht installiert. Excel-Export ist nicht verfügbar.")
        
        # Fortschritt aktualisieren
        update_task_progress(self.request.id, 10, "Exportdaten werden validiert")
        
        # Parameter validieren
        validate_report_parameters(export_data, required_fields=['sheets'])
        
        # Daten extrahieren
        sheets = export_data.get('sheets', [])
        include_charts = export_data.get('include_charts', True)
        include_summary = export_data.get('include_summary', True)
        
        # Ausgabepfad bestimmen
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = os.path.join('reports', 'excel')
            os.makedirs(output_dir, exist_ok=True)
            output_file = os.path.join(output_dir, f"export_{timestamp}.xlsx")
        
        # Sicherstellen, dass das Ausgabeverzeichnis existiert
        os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)
        
        update_task_progress(self.request.id, 20, "Excel-Datei wird vorbereitet")
        
        # Excel-Writer erstellen
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Für jedes Blatt im Export
            for i, sheet_data in enumerate(sheets):
                progress = 20 + int(60 * (i + 1) / len(sheets))
                update_task_progress(self.request.id, progress, f"Blatt {i+1}/{len(sheets)} wird erstellt")
                
                sheet_name = sheet_data.get('name', f'Sheet{i+1}')
                data = sheet_data.get('data', [])
                columns = sheet_data.get('columns', [])
                
                # Daten in DataFrame umwandeln
                if isinstance(data, list) and all(isinstance(item, dict) for item in data):
                    # Liste von Dictionaries
                    df = pd.DataFrame(data)
                elif isinstance(data, list) and all(isinstance(item, list) for item in data):
                    # Liste von Listen
                    df = pd.DataFrame(data, columns=columns if columns else None)
                else:
                    # Ungültiges Datenformat
                    logger.warning(f"Ungültiges Datenformat für Blatt {sheet_name}, überspringe")
                    continue
                
                # DataFrame in Excel schreiben
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Arbeitsblatt formatieren
                worksheet = writer.sheets[sheet_name]
                
                # Überschriften formatieren
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Spaltenbreiten anpassen
                for i, column in enumerate(worksheet.columns):
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(i + 1)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
                # Diagramme hinzufügen, falls gewünscht
                if include_charts and sheet_data.get('charts'):
                    for chart_data in sheet_data.get('charts', []):
                        chart_type = chart_data.get('type', 'bar')
                        chart_title = chart_data.get('title', 'Diagramm')
                        data_cols = chart_data.get('data_cols', [])
                        category_col = chart_data.get('category_col')
                        
                        if not data_cols or not category_col:
                            continue
                        
                        # Diagramm erstellen
                        if chart_type == 'bar':
                            chart = BarChart()
                        elif chart_type == 'line':
                            chart = LineChart()
                        elif chart_type == 'pie':
                            chart = PieChart()
                        else:
                            continue
                        
                        chart.title = chart_title
                        
                        # Datenbereich bestimmen
                        last_row = len(df) + 1  # +1 für die Überschrift
                        
                        # Kategorienspalte finden
                        cat_col_idx = df.columns.get_loc(category_col) + 1  # +1 für Excel-Indizierung
                        cats = Reference(worksheet, min_col=cat_col_idx, min_row=1, max_row=last_row)
                        
                        # Datenspalten finden und hinzufügen
                        for data_col in data_cols:
                            if data_col in df.columns:
                                data_col_idx = df.columns.get_loc(data_col) + 1  # +1 für Excel-Indizierung
                                data = Reference(worksheet, min_col=data_col_idx, min_row=1, max_row=last_row)
                                series = chart.series.append(data)
                                series.title = data_col
                        
                        # Kategorien setzen
                        chart.set_categories(cats)
                        
                        # Diagramm zum Arbeitsblatt hinzufügen
                        worksheet.add_chart(chart, "H2")
            
            # Zusammenfassungsblatt hinzufügen, falls gewünscht
            if include_summary:
                update_task_progress(self.request.id, 80, "Zusammenfassungsblatt wird erstellt")
                
                summary_data = []
                summary_data.append(["Exportzusammenfassung", ""])
                summary_data.append(["Erstellt am", datetime.now().strftime('%d.%m.%Y %H:%M')])
                summary_data.append(["Anzahl der Blätter", len(sheets)])
                
                total_records = 0
                for sheet_data in sheets:
                    sheet_name = sheet_data.get('name', '')
                    data = sheet_data.get('data', [])
                    record_count = len(data)
                    total_records += record_count
                    summary_data.append([f"Datensätze in {sheet_name}", record_count])
                
                summary_data.append(["Gesamtanzahl Datensätze", total_records])
                
                # Zusammenfassungs-DataFrame erstellen und in Excel schreiben
                summary_df = pd.DataFrame(summary_data, columns=["Metrik", "Wert"])
                summary_df.to_excel(writer, sheet_name="Zusammenfassung", index=False)
                
                # Zusammenfassungsblatt formatieren
                summary_sheet = writer.sheets["Zusammenfassung"]
                
                # Überschriften formatieren
                for cell in summary_sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Spaltenbreiten anpassen
                for i, column in enumerate(summary_sheet.columns):
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(i + 1)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2) * 1.2
                    summary_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        update_task_progress(self.request.id, 100, "Excel-Export erfolgreich generiert")
        
        # Ergebnis zurückgeben
        return {
            "status": "success",
            "export_file": output_file,
            "export_size_bytes": os.path.getsize(output_file),
            "sheets_count": len(sheets),
            "generated_at": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Fehler beim Excel-Export: {str(e)}")
        raise

@shared_task(bind=True)
def generate_data_analysis_export(self, analysis_results: Dict[str, Any],
                               include_visualizations: bool = True,
                               output_file: Optional[str] = None) -> Dict[str, Any]:
    """
    Generiert einen Excel-Export für Datenanalyse-Ergebnisse.
    
    Args:
        analysis_results: Dictionary mit Analyseergebnissen
        include_visualizations: Ob Visualisierungen im Export enthalten sein sollen
        output_file: Pfad zur Ausgabedatei (optional)
    
    Returns:
        Dict mit Informationen zum generierten Export
    """
    logger.info("Starte Export von Datenanalyse-Ergebnissen")
    
    try:
        # Daten für den Excel-Export vorbereiten
        export_data = {
            "sheets": [],
            "include_charts": include_visualizations,
            "include_summary": True
        }
        
        # Für jede Metrik ein Blatt erstellen
        for metric_name, metric_data in analysis_results.items():
            # Zeitreihen in ein Blatt umwandeln
            if "time_series" in metric_data:
                time_series = metric_data["time_series"]
                
                # Daten in ein Format umwandeln, das für Excel geeignet ist
                sheet_data = []
                for point in time_series:
                    sheet_data.append({
                        "Datum": point.get("date", ""),
                        "Wert": point.get("value", 0)
                    })
                
                # Diagramm für die Zeitreihe definieren
                charts = []
                if include_visualizations:
                    charts.append({
                        "type": "line",
                        "title": f"{metric_name} Zeitreihe",
                        "category_col": "Datum",
                        "data_cols": ["Wert"]
                    })
                
                # Blatt zum Export hinzufügen
                export_data["sheets"].append({
                    "name": f"{metric_name}_Zeitreihe",
                    "data": sheet_data,
                    "charts": charts
                })
            
            # Statistiken in ein separates Blatt umwandeln
            if "statistics" in metric_data:
                stats = metric_data["statistics"]
                
                # Statistiken in ein Format umwandeln, das für Excel geeignet ist
                stats_data = []
                for stat_name, stat_value in stats.items():
                    stats_data.append({
                        "Statistik": stat_name,
                        "Wert": stat_value
                    })
                
                # Blatt zum Export hinzufügen
                export_data["sheets"].append({
                    "name": f"{metric_name}_Statistik",
                    "data": stats_data,
                    "charts": []
                })
        
        # Excel-Export mit den vorbereiteten Daten generieren
        return generate_excel_export(self, export_data, output_file=output_file)
        
    except Exception as e:
        logger.error(f"Fehler beim Export von Datenanalyse-Ergebnissen: {str(e)}")
        raise

@shared_task(bind=True)
def generate_transaction_report(self, transaction_data: Dict[str, Any],
                             group_by: str = 'day',
                             output_file: Optional[str] = None) -> Dict[str, Any]:
    """
    Generiert einen Excel-Bericht für Transaktionsdaten.
    
    Args:
        transaction_data: Dictionary mit Transaktionsdaten
        group_by: Gruppierung der Daten ('day', 'week', 'month', 'quarter')
        output_file: Pfad zur Ausgabedatei (optional)
    
    Returns:
        Dict mit Informationen zum generierten Bericht
    """
    logger.info(f"Starte Transaktionsbericht mit Gruppierung nach {group_by}")
    
    try:
        # Transaktionsdaten extrahieren
        transactions = transaction_data.get('transactions', [])
        
        if not transactions:
            raise ValueError("Keine Transaktionsdaten vorhanden")
        
        # Daten in DataFrame umwandeln
        df = pd.DataFrame(transactions)
        
        # Datum konvertieren
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'])
        
        # Gruppierung nach Zeitraum
        if group_by == 'day':
            df['group'] = df['date'].dt.date
        elif group_by == 'week':
            df['group'] = df['date'].dt.isocalendar().week
            df['group_name'] = df['date'].dt.strftime('KW %V %Y')
        elif group_by == 'month':
            df['group'] = df['date'].dt.month
            df['group_name'] = df['date'].dt.strftime('%B %Y')
        elif group_by == 'quarter':
            df['group'] = df['date'].dt.quarter
            df['group_name'] = 'Q' + df['date'].dt.quarter.astype(str) + ' ' + df['date'].dt.year.astype(str)
        
        # Gruppierte Zusammenfassung erstellen
        summary = df.groupby('group').agg({
            'amount': ['sum', 'mean', 'count'],
            'group_name': 'first' if 'group_name' in df.columns else None
        }).reset_index()
        
        # Spalten umbenennen
        summary.columns = ['group', 'total_amount', 'avg_amount', 'transaction_count', 
                         'period' if 'group_name' in df.columns else None]
        
        # Wenn group_name nicht existiert, group als period verwenden
        if 'group_name' not in df.columns:
            summary = summary.drop(columns=['period'])
            summary['period'] = summary['group'].astype(str)
        
        # Nach Kategorien gruppieren, falls vorhanden
        category_summary = None
        if 'category' in df.columns:
            category_summary = df.groupby('category').agg({
                'amount': ['sum', 'mean', 'count']
            }).reset_index()
            
            category_summary.columns = ['category', 'total_amount', 'avg_amount', 'transaction_count']
        
        # Daten für den Excel-Export vorbereiten
        export_data = {
            "sheets": [
                {
                    "name": "Transaktionen",
                    "data": transactions,
                    "charts": []
                },
                {
                    "name": f"Zusammenfassung_{group_by}",
                    "data": summary.to_dict('records'),
                    "charts": [
                        {
                            "type": "bar",
                            "title": f"Gesamtbetrag nach {group_by}",
                            "category_col": "period",
                            "data_cols": ["total_amount"]
                        },
                        {
                            "type": "line",
                            "title": f"Transaktionsanzahl nach {group_by}",
                            "category_col": "period",
                            "data_cols": ["transaction_count"]
                        }
                    ]
                }
            ],
            "include_charts": True,
            "include_summary": True
        }
        
        # Kategorie-Zusammenfassung hinzufügen, falls vorhanden
        if category_summary is not None:
            export_data["sheets"].append({
                "name": "Kategorien",
                "data": category_summary.to_dict('records'),
                "charts": [
                    {
                        "type": "pie",
                        "title": "Verteilung nach Kategorien",
                        "category_col": "category",
                        "data_cols": ["total_amount"]
                    }
                ]
            })
        
        # Excel-Export mit den vorbereiteten Daten generieren
        return generate_excel_export(self, export_data, output_file=output_file)
        
    except Exception as e:
        logger.error(f"Fehler bei der Generierung des Transaktionsberichts: {str(e)}")
        raise 