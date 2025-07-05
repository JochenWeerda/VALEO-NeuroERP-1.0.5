"""
Tests für den Excel-Export-Service
"""

import pytest
from datetime import datetime
from fastapi.testclient import TestClient
from unittest.mock import Mock, patch
import openpyxl
from backend.services.excel_export_service import (
    ExportConfig,
    ExportJob,
    router,
    process_export,
    load_table_data,
    format_worksheet,
    create_pivot_tables
)

# Test Data
SAMPLE_TABLE_DATA = [
    {"id": 1, "name": "Test 1", "value": 100.50, "date": "2025-01-01"},
    {"id": 2, "name": "Test 2", "value": 200.75, "date": "2025-01-02"},
]

@pytest.fixture
def test_client():
    """Test client für FastAPI"""
    from fastapi import FastAPI
    app = FastAPI()
    app.include_router(router)
    return TestClient(app)

@pytest.fixture
def mock_redis():
    """Mock für Redis"""
    with patch("backend.services.excel_export_service.Redis") as mock:
        yield mock

@pytest.fixture
def sample_config():
    """Test-Konfiguration für Excel-Export"""
    return ExportConfig(
        table_id="test_table",
        format="xlsx",
        include_styles=True,
        include_pivot=True,
        date_format="DD.MM.YYYY",
        number_format="#,##0.00",
        sheet_name="Test"
    )

@pytest.fixture
def sample_job(sample_config):
    """Test-Job für Excel-Export"""
    return ExportJob(
        id="test_export_20250101_120000",
        user_id="test_user",
        config=sample_config,
        status="pending",
        progress=0.0,
        created_at=datetime(2025, 1, 1, 12, 0, 0)
    )

class TestExcelExportService:
    """Testfälle für Excel-Export-Service"""
    
    async def test_create_excel_export(self, test_client, mock_redis, sample_config):
        """Test: Export-Job erstellen"""
        response = test_client.post(
            "/api/v1/export/excel",
            json=sample_config.dict(),
            headers={"X-User-ID": "test_user"}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "pending"
        assert data["progress"] == 0.0
        assert "id" in data
        
    async def test_get_export_status(self, test_client, mock_redis, sample_job):
        """Test: Export-Status abrufen"""
        mock_redis.return_value.get.return_value = sample_job.json()
        
        response = test_client.get(f"/api/v1/export/{sample_job.id}/status")
        
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == sample_job.id
        assert data["status"] == "pending"
        
    async def test_get_export_status_not_found(self, test_client, mock_redis):
        """Test: Export-Status für nicht existierenden Job"""
        mock_redis.return_value.get.return_value = None
        
        response = test_client.get("/api/v1/export/nonexistent/status")
        
        assert response.status_code == 404
        
    @pytest.mark.asyncio
    async def test_process_export(self, sample_job, mock_redis):
        """Test: Export-Verarbeitung"""
        with patch("backend.services.excel_export_service.load_table_data") as mock_load:
            mock_load.return_value = SAMPLE_TABLE_DATA
            
            await process_export(sample_job)
            
            assert sample_job.status == "completed"
            assert sample_job.progress == 100.0
            assert sample_job.file_path.endswith(".xlsx")
            
    @pytest.mark.asyncio
    async def test_load_table_data(self):
        """Test: Tabellendaten laden"""
        data = await load_table_data("test_table")
        
        assert isinstance(data, list)
        for item in data:
            assert isinstance(item, dict)
            assert "id" in item
            
    @pytest.mark.asyncio
    async def test_format_worksheet(self, sample_config):
        """Test: Worksheet formatieren"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        await format_worksheet(ws, SAMPLE_TABLE_DATA, sample_config)
        
        # Prüfen der Formatierung
        assert ws.cell(1, 1).value == "ID"  # Header
        assert ws.cell(2, 1).value == 1     # Erste Zeile
        assert ws.cell(2, 2).value == "Test 1"
        
    @pytest.mark.asyncio
    async def test_create_pivot_tables(self):
        """Test: Pivot-Tabellen erstellen"""
        wb = openpyxl.Workbook()
        
        await create_pivot_tables(wb, SAMPLE_TABLE_DATA)
        
        # Prüfen der Pivot-Tabelle
        assert len(wb.sheetnames) > 1  # Mindestens ein zusätzliches Sheet
        pivot_sheet = wb["Pivot"]
        assert pivot_sheet is not None 